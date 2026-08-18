import random
import re
import json
import openpyxl
import razorpay
from decimal import Decimal
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.mail import send_mail
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.utils import timezone
from datetime import timedelta
from django.http import JsonResponse
from django.contrib.sessions.models import Session
from django.urls import reverse
from django.core.paginator import Paginator
from django.db.models import Q

from .models import (
    CustomUser, TeacherProfile, StudentProfile, MockTest, Question, TestAttempt, StudentAnswer,
    PaymentTransaction, WalletTransaction, WithdrawalRequest,
    ExamCategory, SubjectCategory, ChapterCategory, TestQuestionMapping
)

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))


# ==========================================
# 1. AUTH & HOME VIEWS
# ==========================================
# ==========================================
# 1. AUTH & HOME VIEWS
# ==========================================
def home(request):
    if request.user.is_authenticated:

        # 🚀 NAYA: Auto-Logout Logic for Demo Users
        # अगर यूज़र का मोबाइल नंबर 'DEMO' से शुरू होता है, तो उसे चुपचाप लॉगआउट कर दें
        if request.user.mobile_number and request.user.mobile_number.startswith('DEMO'):
            logout(request)
            return redirect('home')

        # अगर असली एडमिन है, तो बैंक में भेजें
        if request.user.is_superuser or request.user.is_staff:
            return redirect('upload_global_bank')

        # अगर असली टीचर है, तो डैशबोर्ड में भेजें
        if request.user.role == 'TEACHER':
            return redirect('teacher_dashboard')

        # अगर असली स्टूडेंट है, तो उसका डैशबोर्ड दिखाएं
        if request.user.role == 'STUDENT':
            try:
                student = request.user.student_profile
                return render(request, 'portal/home.html', {'has_plan': student.has_active_plan()})
            except StudentProfile.DoesNotExist:
                pass

    return render(request, 'portal/home.html', {})

# ==========================================
# 🚀 NAYA: QUICK JOIN / DEMO TEST LOGIC
# ==========================================
# ==========================================
# 🚀 QUICK JOIN / DEMO TEST LOGIC
# ==========================================
def quick_join_test(request):
    if request.method == 'POST':
        student_name = request.POST.get('student_name', '').strip()
        coaching_code = request.POST.get('coaching_code', '').strip()
        test_code = request.POST.get('test_code', '').strip().upper()

        try:
            teacher = TeacherProfile.objects.get(coaching_code=coaching_code)
        except TeacherProfile.DoesNotExist:
            messages.error(request, '❌ गलत कोचिंग कोड! कृपया सही कोड दर्ज करें।')
            return redirect('quick_join_test')  # 🚀 NAYA: वापस इसी पेज पर भेजें

        if not teacher.is_demo_active():
            messages.error(request,
                           '⏳ इस कोचिंग का फ्री डेमो पीरियड समाप्त हो गया है। कृपया अपने मोबाइल नंबर से रजिस्टर करके टेस्ट दें।')
            return redirect('home')

        try:
            test = MockTest.objects.get(test_code=test_code, teacher=teacher, status='PUBLISHED')
        except MockTest.DoesNotExist:
            messages.error(request, '❌ गलत टेस्ट कोड या यह टेस्ट अभी लाइव नहीं है!')
            return redirect('quick_join_test')  # 🚀 NAYA: वापस इसी पेज पर भेजें

        guest_mobile = f"DEMO{random.randint(100000, 999999)}"
        guest_email = f"demo_{guest_mobile}@quickjoin.com"

        user = CustomUser.objects.create_user(
            mobile_number=guest_mobile, email=guest_email, password='guestpassword123',
            full_name=student_name, role='STUDENT'
        )

        StudentProfile.objects.create(user=user, enrolled_coaching=teacher, display_name=f"{student_name} (Demo)")

        login(request, user)
        user.last_session_key = request.session.session_key
        user.save()

        messages.success(request, f'🎉 स्वागत है {student_name}! आपका फ्री डेमो टेस्ट शुरू हो रहा है।')
        return redirect('test_instructions', test_id=test.id)

    # 🚀 NAYA: अगर POST नहीं है (यानी यूजर ने बटन पर क्लिक किया है), तो यह नया पेज खोलें
    return render(request, 'portal/quick_join.html')

def signup_view(request):
    if request.method == 'POST':
        full_name = request.POST.get('full_name')
        mobile_number = request.POST.get('mobile_number')
        email = request.POST.get('email')
        coaching_code = request.POST.get('coaching_code')
        password = request.POST.get('password')

        if CustomUser.objects.filter(mobile_number=mobile_number).exists():
            messages.error(request, 'यह मोबाइल नंबर पहले से रजिस्टर है।')
            return redirect('signup')

        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, 'यह ईमेल पहले से रजिस्टर है।')
            return redirect('signup')

        try:
            teacher = TeacherProfile.objects.get(coaching_code=coaching_code)
        except TeacherProfile.DoesNotExist:
            messages.error(request, 'कोचिंग कोड गलत है! कृपया सही कोड डालें।')
            return redirect('signup')

        otp = str(random.randint(100000, 999999))
        request.session['temp_user'] = {
            'full_name': full_name, 'mobile_number': mobile_number,
            'email': email, 'coaching_code': coaching_code, 'password': password
        }
        request.session['otp'] = otp

        # 🚀 NAYA: HTML Email Template for Signup
        current_time = timezone.localtime(timezone.now()).strftime("%d %b %Y, %I:%M %p")
        subject = f'Mock Test Portal OTP: {otp}'
        plain_message = f'आपका रजिस्ट्रेशन OTP है: {otp}'

        html_message = f"""
        <div style="font-family: Arial, sans-serif; background-color: #f4ebdd; padding: 30px 10px;">
            <div style="max-width: 450px; margin: auto; background-color: #ffffff; padding: 30px; border-radius: 12px; border: 1px solid #ddd; box-shadow: 0 4px 10px rgba(0,0,0,0.05);">
                <h2 style="text-align: center; color: #ffc107; margin-top: 0; font-size: 24px;">Mock Test Portal 📚</h2>
                <p style="color: #333; font-size: 16px;">नमस्ते <b>{full_name}</b>,</p>
                <p style="color: #555; font-size: 15px; line-height: 1.5;">Mock Test Portal में आपका स्वागत है! आपके अकाउंट को वेरीफाई करने का OTP नीचे दिया गया है:</p>

                <div style="background-color: #4ade80; color: #000; font-size: 36px; font-weight: bold; letter-spacing: 6px; text-align: center; padding: 15px; border-radius: 8px; margin: 25px 0;">
                    {otp}
                </div>

                <p style="font-size: 13px; color: #666; line-height: 1.6;">यह OTP <b>{current_time}</b> पर भेजा गया है और केवल <b>10 मिनट</b> के लिए मान्य है। कृपया इसे किसी के साथ शेयर न करें।</p>

                <hr style="border: none; border-top: 1px solid #eee; margin: 25px 0;">
                <p style="font-size: 14px; color: #888; text-align: center;">धन्यवाद,<br><b style="color: #333;">Mock Test Portal Team</b></p>
            </div>
        </div>
        """

        # 🛡️ 🚀 NAYA: Try-Except Block (Crash Protection)
        try:
            send_mail(
                subject,
                plain_message,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
                html_message=html_message  # 🚀 HTML Message यहाँ से भेजा जाएगा
            )
            return redirect('verify_otp')

        except Exception as e:
            # 🛑 अगर ईमेल फेल हो जाए, तो साइट क्रैश नहीं होगी, बल्कि असल एरर स्क्रीन पर प्रिंट कर देगी!
            messages.error(request, f'⚠️ ईमेल एरर: {str(e)}')
            return redirect('signup')

    return render(request, 'portal/signup.html')


def verify_otp_view(request):
    if request.method == 'POST':
        entered_otp = request.POST.get('otp')
        saved_otp = request.session.get('otp')
        user_data = request.session.get('temp_user')

        if entered_otp == saved_otp and user_data:
            user = CustomUser.objects.create_user(
                mobile_number=user_data['mobile_number'], email=user_data['email'],
                password=user_data['password'], full_name=user_data['full_name'], role='STUDENT'
            )
            teacher = TeacherProfile.objects.get(coaching_code=user_data['coaching_code'])
            StudentProfile.objects.create(user=user, enrolled_coaching=teacher)
            del request.session['temp_user']
            del request.session['otp']
            messages.success(request, 'अकाउंट सफलतापूर्वक बन गया! अब आप लॉगिन कर सकते हैं।')
            return redirect('login')
        else:
            messages.error(request, 'OTP गलत है! कृपया दोबारा प्रयास करें।')
    return render(request, 'portal/verify_otp.html')


def login_view(request):
    if request.method == 'POST':
        mobile_number = request.POST.get('mobile_number')
        password = request.POST.get('password')
        user = authenticate(request, mobile_number=mobile_number, password=password)

        if user is not None:
            if user.last_session_key:
                if Session.objects.filter(session_key=user.last_session_key).exists():
                    request.session['pre_auth_user_id'] = user.id
                    return redirect('confirm_device_login')

            login(request, user)
            user.last_session_key = request.session.session_key
            user.save()

            if user.is_superuser or user.is_staff:
                return redirect('upload_global_bank')
            elif user.role == 'TEACHER':
                return redirect('teacher_dashboard')
            else:
                return redirect('home')
        else:
            messages.error(request, 'मोबाइल नंबर या पासवर्ड गलत है!')
    return render(request, 'portal/login.html')


def confirm_device_login(request):
    user_id = request.session.get('pre_auth_user_id')
    if not user_id:
        return redirect('login')

    user = CustomUser.objects.get(id=user_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'remain':
            if 'pre_auth_user_id' in request.session:
                del request.session['pre_auth_user_id']
            messages.info(request, "लॉगिन रद्द किया गया। आप अपने पिछले डिवाइस पर लॉगिन बने हुए हैं।")
            return redirect('login')

        elif action == 'kickout':
            old_session_key = user.last_session_key
            current_session_key = request.session.session_key
            if old_session_key and old_session_key != current_session_key:
                Session.objects.filter(session_key=old_session_key).delete()

            login(request, user)
            user.last_session_key = request.session.session_key
            user.save()

            if 'pre_auth_user_id' in request.session:
                del request.session['pre_auth_user_id']

            if user.is_superuser or user.is_staff:
                return redirect('upload_global_bank')
            elif user.role == 'TEACHER':
                return redirect('teacher_dashboard')
            else:
                return redirect('home')
    return render(request, 'portal/confirm_device.html', {'target_user': user})


def logout_view(request):
    logout(request)
    return redirect('login')


# ==========================================
# 2. TEACHER DASHBOARD & TEST MANAGEMENT
# ==========================================
@login_required
def teacher_dashboard(request):
    if request.user.role != 'TEACHER': return redirect('home')
    teacher_profile = request.user.teacher_profile
    form_data = {}

    if request.method == 'POST' and 'create_test' in request.POST:
        form_data = request.POST
        title, test_code, total_time, correct_marks, negative_marks = request.POST.get('title'), request.POST.get(
            'test_code').upper().strip(), request.POST.get('total_time'), request.POST.get(
            'correct_marks'), request.POST.get('negative_marks')
        scheduled_time = request.POST.get('scheduled_time') or None

        if MockTest.objects.filter(test_code=test_code).exists():
            messages.error(request, f'टेस्ट कोड "{test_code}" पहले से मौजूद है! कृपया नया कोड डालें।')
        else:
            MockTest.objects.create(
                teacher=teacher_profile, title=title, test_code=test_code, total_time=total_time,
                correct_marks=correct_marks, negative_marks=negative_marks, scheduled_time=scheduled_time,
                status='DRAFT'
            )
            messages.success(request, f'टेस्ट "{title}" बन गया है! अब Excel अपलोड करें।')
            return redirect('teacher_dashboard')

    for test in MockTest.objects.filter(teacher=teacher_profile, status='PUBLISHED'):
        if test.scheduled_time and test.total_time:
            if timezone.now() > test.scheduled_time + timedelta(minutes=int(test.total_time)):
                test.status = 'DRAFT'
                test.save()

    tests = MockTest.objects.filter(teacher=teacher_profile).order_by('-created_at')
    return render(request, 'portal/teacher_dashboard.html', {'tests': tests, 'form_data': form_data})


@login_required
def toggle_publish_status(request, test_id):
    if request.user.role != 'TEACHER': return redirect('home')
    try:
        test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
        if test.status == 'DRAFT':
            if test.mapped_questions.count() == 0:
                messages.error(request, 'एरर: बिना प्रश्न जोड़े आप टेस्ट को पब्लिश नहीं कर सकते!')
                return redirect('teacher_dashboard')
            test.status = 'PUBLISHED'
            messages.success(request, f'टेस्ट "{test.title}" अब लाइव हो गया है!')
        else:
            test.status = 'DRAFT'
            messages.success(request, f'टेस्ट "{test.title}" अब ड्राफ्ट मोड में आ गया है।')
        test.save()
    except MockTest.DoesNotExist:
        pass
    return redirect('teacher_dashboard')


@login_required
def edit_schedule(request, test_id):
    if request.user.role != 'TEACHER': return redirect('home')
    try:
        test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
        if request.method == 'POST':
            new_title = request.POST.get('new_title')
            new_code = request.POST.get('new_test_code').upper().strip()
            new_scheduled_time = request.POST.get('new_scheduled_time') or None

            if new_code != test.test_code:
                if MockTest.objects.filter(test_code=new_code).exists():
                    messages.error(request, f'टेस्ट कोड "{new_code}" पहले से इस्तेमाल में है!')
                    return redirect('teacher_dashboard')
                with transaction.atomic():
                    new_test = MockTest.objects.create(
                        teacher=test.teacher, title=new_title, test_code=new_code,
                        total_time=request.POST.get('new_total_time'),
                        correct_marks=request.POST.get('new_correct_marks'),
                        negative_marks=request.POST.get('new_negative_marks'), scheduled_time=new_scheduled_time,
                        status='DRAFT'
                    )
                    for mapping in test.mapped_questions.all():
                        TestQuestionMapping.objects.create(test=new_test, question=mapping.question,
                                                           order=mapping.order)
                messages.success(request, f'नया टेस्ट "{new_test.title}" सफलता से कॉपी हो गया!')
            else:
                test.title, test.total_time, test.correct_marks, test.negative_marks, test.scheduled_time = new_title, request.POST.get(
                    'new_total_time'), request.POST.get('new_correct_marks'), request.POST.get(
                    'new_negative_marks'), new_scheduled_time
                test.save()
                messages.success(request, 'टेस्ट सेटिंग्स सफलतापूर्वक अपडेट हो गईं!')
    except Exception as e:
        messages.error(request, 'कुछ गलती हुई।')
    return redirect('teacher_dashboard')


@login_required
def delete_test(request, test_id):
    if request.user.role != 'TEACHER': return redirect('home')
    try:
        test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
        test.delete()
        messages.success(request, f'टेस्ट हमेशा के लिए डिलीट कर दिया गया है!')
    except MockTest.DoesNotExist:
        pass
    return redirect('teacher_dashboard')


@login_required
def preview_test(request, test_id):
    test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
    questions = [m.question for m in test.mapped_questions.select_related('question').all()]
    return render(request, 'portal/preview_test.html', {'test': test, 'questions': questions})


@login_required
def upload_excel(request, test_id):
    test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            count = 0
            current_order = test.mapped_questions.count()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                q = Question.objects.create(
                    assigned_coaching=request.user.teacher_profile,
                    question_text=str(row[0]), option_a=str(row[1]), option_b=str(row[2]),
                    option_c=str(row[3]), option_d=str(row[4]), correct_answer=str(row[5]).strip().upper(),
                    explanation=str(row[6]) if len(row) > 6 and row[6] else ""
                )
                count += 1
                TestQuestionMapping.objects.create(test=test, question=q, order=current_order + count)
            messages.success(request, f'✅ सफलता! आपके {count} प्रश्न सुरक्षित रूप से अपलोड हो गए हैं।')
        except Exception as e:
            messages.error(request, f'एक्सेल फाइल पढ़ने में एरर: {str(e)}')
        return redirect('teacher_dashboard')
    return render(request, 'portal/upload_excel.html', {'test': test})


@login_required
def delete_questions(request, test_id):
    test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
    test.mapped_questions.all().delete()
    messages.success(request, '🗑️ सारे प्रश्न इस टेस्ट से हटा दिए गए हैं।')
    return redirect('teacher_dashboard')


# ==========================================
# 3. SMART CHOOSE QUESTIONS & BANK LOGIC
# ==========================================
# ==========================================
# 3. SMART CHOOSE QUESTIONS & BANK LOGIC
# ==========================================
@login_required
def choose_questions(request):
    # 🚀 FIX 1: एडमिन को भी पेज देखने की परमिशन दें
    if request.user.role != 'TEACHER' and not request.user.is_superuser:
        return redirect('home')

    if request.GET.get('ajax') == '1':
        exam_id = request.GET.get('exam', 'ALL')
        sub_id = request.GET.get('subject', 'ALL')
        chap_id = request.GET.get('chapter', 'ALL')
        search_text = request.GET.get('search', '').strip()
        bank_type = request.GET.get('bank_type', 'GLOBAL')

        if bank_type == 'PRIVATE':
            # 🚀 FIX 2: एडमिन का प्राइवेट बैंक नहीं होता, इसलिए उसे खाली (None) दिखाएं
            if request.user.is_superuser:
                questions = Question.objects.none()
            else:
                questions = Question.objects.filter(assigned_coaching=request.user.teacher_profile).order_by(
                    '-created_at')
        else:
            questions = Question.objects.filter(assigned_coaching__isnull=True).order_by('-created_at')

        if chap_id != 'ALL':
            questions = questions.filter(chapter_id=chap_id)
        elif sub_id != 'ALL':
            questions = questions.filter(chapter__subject_id=sub_id)
        elif exam_id != 'ALL':
            questions = questions.filter(chapter__subject__exam_id=exam_id)

        if len(search_text) >= 3:
            questions = questions.filter(Q(question_text__icontains=search_text) | Q(option_a__icontains=search_text))

        paginator = Paginator(questions, 20)
        page_obj = paginator.get_page(request.GET.get('page', 1))

        data = []
        for q in page_obj:
            data.append({
                'id': q.id,
                'question_text': q.question_text,
                'exam': q.chapter.subject.exam.name if q.chapter else '',
                'subject': q.chapter.subject.name if q.chapter else '',
                'chapter': q.chapter.name if q.chapter else '',
                'option_a': q.option_a, 'option_b': q.option_b, 'option_c': q.option_c, 'option_d': q.option_d,
                'correct_answer': q.correct_answer,
            })
        return JsonResponse({'questions': data, 'has_next': page_obj.has_next(), 'total_results': paginator.count})

    # 🚀 FIX 3: एडमिन के पास teacher_profile नहीं होता, इसलिए एरर से बचाएं
    if request.user.is_superuser:
        private_exams = []
        draft_tests = []
    else:
        teacher = request.user.teacher_profile
        private_exams = ExamCategory.objects.filter(created_by=teacher)
        draft_tests = MockTest.objects.filter(teacher=teacher, status='DRAFT')

    global_exams = ExamCategory.objects.filter(created_by__isnull=True)

    return render(request, 'portal/choose_questions.html', {
        'private_exams': private_exams,
        'global_exams': global_exams,
        'draft_tests': draft_tests
    })

@login_required
def api_add_bank_questions(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            test_id = data.get('test_id')
            q_ids = data.get('q_ids', [])

            if not q_ids: return JsonResponse({'status': 'error', 'message': 'कोई प्रश्न सेलेक्ट नहीं किया गया है!'})

            if test_id:
                test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
            else:
                code = data.get('test_code').upper().strip()
                if MockTest.objects.filter(test_code=code).exists():
                    return JsonResponse({'status': 'error', 'message': f'टेस्ट कोड "{code}" पहले से मौजूद है!'})
                test = MockTest.objects.create(
                    teacher=request.user.teacher_profile, title=data.get('title'), test_code=code,
                    total_time=data.get('total_time', 10), correct_marks=data.get('correct_marks', 4),
                    negative_marks=data.get('negative_marks', 1), status='DRAFT'
                )

            current_max_order = TestQuestionMapping.objects.filter(test=test).count()
            added_count = 0
            for i, q in enumerate(Question.objects.filter(id__in=q_ids), start=1):
                if not TestQuestionMapping.objects.filter(test=test, question=q).exists():
                    TestQuestionMapping.objects.create(test=test, question=q, order=current_max_order + i)
                    added_count += 1
            return JsonResponse(
                {'status': 'success', 'message': f'🎉 {added_count} प्रश्न सफलतापूर्वक "{test.title}" में जुड़ गए हैं!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid Request'})


# ==========================================
# 4. STUDENT EXAM ENGINE
# ==========================================
@login_required
def start_test(request):
    if request.user.role != 'STUDENT': return redirect('teacher_dashboard')
    if not request.user.student_profile.has_active_plan(): return redirect('initiate_payment')

    if request.method == 'POST':
        test_code = request.POST.get('test_code').strip().upper()
        try:
            test = MockTest.objects.get(test_code=test_code, status='PUBLISHED')
        except MockTest.DoesNotExist:
            messages.error(request, 'गलत टेस्ट कोड! या टेस्ट अभी लाइव नहीं हुआ है।')
            return redirect('home')

        if TestAttempt.objects.filter(student=request.user.student_profile, test=test, is_completed=True).exists():
            messages.error(request, 'आप यह टेस्ट पहले ही दे चुके हैं!')
            return redirect('home')
        return redirect('test_instructions', test_id=test.id)
    return redirect('home')


@login_required
def test_instructions(request, test_id):
    test = MockTest.objects.get(id=test_id)
    attempt, created = TestAttempt.objects.get_or_create(student=request.user.student_profile, test=test)
    is_early, formatted_time = False, ""
    if test.scheduled_time and timezone.now() < test.scheduled_time:
        is_early = True
        formatted_time = timezone.localtime(test.scheduled_time).strftime("%d %b, %I:%M %p")

    if request.method == 'POST':
        if is_early:
            messages.error(request, f'⏳ कृपया प्रतीक्षा करें! यह टेस्ट {formatted_time} पर शुरू होगा।')
            return redirect('test_instructions', test_id=test.id)
        return redirect('take_test', attempt_id=attempt.id)
    return render(request, 'portal/instructions.html',
                  {'test': test, 'is_early': is_early, 'formatted_time': formatted_time})


@login_required
def take_test(request, attempt_id):
    attempt = TestAttempt.objects.get(id=attempt_id, student=request.user.student_profile)
    if attempt.is_completed: return redirect('test_result', attempt_id=attempt.id)

    test = attempt.test
    questions = [m.question for m in test.mapped_questions.select_related('question').all()]

    if request.method == 'POST':
        total_score = Decimal('0.00')
        for q in questions:
            selected = request.POST.get(f'question_{q.id}')
            StudentAnswer.objects.create(attempt=attempt, question=q, selected_option=selected)
            if selected:
                total_score += test.correct_marks if selected == q.correct_answer else -test.negative_marks
        attempt.score, attempt.is_completed = total_score, True
        attempt.save()
        return redirect('test_result', attempt_id=attempt.id)

    remaining_seconds = test.total_time * 60
    is_strict = False
    if test.scheduled_time:
        is_strict = True
        now = timezone.now()
        end_time = test.scheduled_time + timedelta(minutes=test.total_time)
        if now >= end_time:
            attempt.is_completed = True
            attempt.save()
            return redirect('test_result', attempt_id=attempt.id)
        remaining_seconds = int((end_time - now).total_seconds())

    return render(request, 'portal/take_test.html', {
        'attempt': attempt, 'test': test, 'questions': questions,
        'coaching_name': test.teacher.user.full_name,
        'student_display_name': f"{request.user.full_name} ({str(request.user.mobile_number)[-4:]})",
        'is_strict': is_strict, 'remaining_seconds': remaining_seconds
    })


@login_required
def test_result(request, attempt_id):
    if request.user.role != 'STUDENT': return redirect('home')
    try:
        attempt = TestAttempt.objects.get(id=attempt_id, student=request.user.student_profile)
        test = attempt.test
        if test.scheduled_time and test.total_time:
            end_time = test.scheduled_time + timedelta(minutes=test.total_time)
            if timezone.now() < end_time:
                return render(request, 'portal/waiting_lounge.html',
                              {'test': test, 'remaining_seconds': int((end_time - timezone.now()).total_seconds()),
                               'attempt': attempt})

        student_answers = attempt.answers.all()
        total_questions = test.mapped_questions.count()
        correct_answers, wrong_answers, answer_data = 0, 0, []

        # लूप चलाकर पहले सही और गलत जवाब गिनें
        for mapping in test.mapped_questions.select_related('question').all():
            q = mapping.question
            ans = student_answers.filter(question=q).first()
            selected = ans.selected_option if ans else None
            is_correct = False

            if selected:  # अगर छात्र ने कोई ऑप्शन चुना है
                if selected == q.correct_answer:
                    correct_answers += 1
                    is_correct = True
                else:
                    wrong_answers += 1

            answer_data.append({'question': q, 'selected': selected, 'is_correct': is_correct})

        # 🚀 NAYA: 100% सटीक गणित (अब यह कभी माइनस में नहीं जाएगा)
        attempted_questions = correct_answers + wrong_answers
        skipped_questions = total_questions - attempted_questions

        # स्कोर कैलकुलेट करें
        total_score = (Decimal(correct_answers) * test.correct_marks) - (Decimal(wrong_answers) * test.negative_marks)
        if attempt.score != total_score:
            attempt.score = total_score
            attempt.save(update_fields=['score'])

        return render(request, 'portal/test_result.html', {
            'test': test,
            'attempt': attempt,
            'total_questions': total_questions,
            'attempted': attempted_questions,
            'unattempted': skipped_questions,  # 🚀 यह वैल्यू अब एकदम सही जाएगी
            'correct_answers': correct_answers,
            'wrong_answers': wrong_answers,
            'total_score': total_score,
            'answer_data': answer_data,
            'coaching_name': test.teacher.user.full_name,
            'student_display_name': f"{request.user.full_name} ({str(request.user.mobile_number)[-4:]})"
        })
    except TestAttempt.DoesNotExist:
        return redirect('home')


@login_required
def test_scoreboard(request, test_id):
    test = MockTest.objects.get(id=test_id)
    if test.scheduled_time and test.total_time:
        if timezone.now() >= test.scheduled_time + timedelta(minutes=test.total_time):
            for p_attempt in TestAttempt.objects.filter(test=test, is_completed=False):
                correct = sum(
                    1 for ans in p_attempt.answers.all() if ans.selected_option == ans.question.correct_answer)
                wrong = sum(1 for ans in p_attempt.answers.all() if
                            ans.selected_option and ans.selected_option != ans.question.correct_answer)
                p_attempt.score, p_attempt.is_completed = (Decimal(correct) * test.correct_marks) - (
                        Decimal(wrong) * test.negative_marks), True
                p_attempt.save()

    attempts = TestAttempt.objects.filter(test=test, is_completed=True)
    questions = [m.question for m in test.mapped_questions.select_related('question').all()]

    student_data = []
    for attempt in attempts:
        ans_dict = {ans.question_id: ans.selected_option for ans in attempt.answers.all()}
        score, q_answers = Decimal('0.00'), {}
        for q in questions:
            selected, is_correct, opt_text = ans_dict.get(q.id), None, ""
            if selected:
                opt_text = f"{selected}. {getattr(q, f'option_{selected.lower()}')}"
                if selected == q.correct_answer:
                    score, is_correct = score + test.correct_marks, True
                else:
                    score -= test.negative_marks;
                    is_correct = False
            q_answers[q.id] = {'text': opt_text, 'is_correct': is_correct, 'selected_letter': selected}
        student_data.append(
            {'name': f"{attempt.student.user.full_name} ({str(attempt.student.user.mobile_number)[-4:]})",
             'score': score, 'answers': q_answers})

    student_data.sort(key=lambda x: x['score'], reverse=True)
    current_rank, previous_score = 1, None
    for student in student_data:
        if previous_score is not None and student['score'] < previous_score: current_rank += 1
        student['rank'], previous_score = current_rank, student['score']
        student['name'] = f"Rank {student['rank']} | {student['name']}"

    rows = []
    for index, q in enumerate(questions, 1):
        row_cells = [dict(student['answers'][q.id], student_name=student['name']) for student in student_data]
        rows.append({'q_number': index, 'q_text': re.sub(r'^Row\s*\d+:\s*', '', q.question_text, flags=re.IGNORECASE),
                     'opt_a': q.option_a, 'opt_b': q.option_b, 'opt_c': q.option_c, 'opt_d': q.option_d,
                     'correct_ans': q.correct_answer, 'cells': row_cells})

    return render(request, 'portal/scoreboard.html', {'test': test, 'student_data': student_data, 'rows': rows,
                                                      'coaching_name': test.teacher.user.full_name})


@login_required
def test_celebration(request, test_id):
    test = MockTest.objects.get(id=test_id)
    attempts = TestAttempt.objects.filter(test=test, is_completed=True).order_by('-score')
    winners, highest_score = [], Decimal('0.00')
    if attempts.exists():
        highest_score = attempts.first().score
        winners = attempts.filter(score=highest_score)
    return render(request, 'portal/celebration.html',
                  {'test': test, 'winners': winners, 'highest_score': highest_score})


@login_required
def test_answer_key(request, test_id):
    test = MockTest.objects.get(id=test_id)
    questions = [m.question for m in test.mapped_questions.select_related('question').all()]
    attempts = TestAttempt.objects.filter(test=test, is_completed=True).order_by('-score')

    current_rank, previous_score, student_attempt = 1, None, None
    for attempt in attempts:
        if previous_score is not None and attempt.score < previous_score: current_rank += 1
        attempt.rank, previous_score = current_rank, attempt.score
        if request.user.role == 'STUDENT' and attempt.student == request.user.student_profile: student_attempt = attempt

    return render(request, 'portal/answer_key.html',
                  {'test': test, 'questions': questions, 'student_attempt': student_attempt,
                   'coaching_name': test.teacher.user.full_name})


# ==========================================
# 5. LIVE MONITORING & AJAX ACTIONS
# ==========================================
@login_required
def live_test_monitor(request, test_id):
    if request.user.role != 'TEACHER': return redirect('home')
    test = MockTest.objects.get(id=test_id, teacher=request.user.teacher_profile)
    questions = [m.question for m in test.mapped_questions.select_related('question').all()]
    return render(request, 'portal/live_monitor.html', {'test': test, 'questions': questions})


@login_required
def api_live_data(request, test_id):
    if request.user.role != 'TEACHER': return JsonResponse({'error': 'Unauthorized'}, status=403)
    test = MockTest.objects.get(id=test_id)
    attempts = TestAttempt.objects.filter(test=test)
    students_data = []

    for attempt in attempts:
        students_data.append({
            'name': f"{attempt.student.user.full_name} ({str(attempt.student.user.mobile_number)[-4:]})",
            'status': "Submitted" if attempt.is_completed else "Solving...",
            'score': attempt.score if attempt.is_completed else "-", 'color': 'green'
        })

    pending_students = StudentProfile.objects.filter(enrolled_coaching=test.teacher).exclude(
        id__in=attempts.values_list('student_id', flat=True))
    for st in pending_students:
        students_data.append(
            {'name': f"{st.user.full_name} ({str(st.user.mobile_number)[-4:]})", 'status': "Pending", 'score': "-",
             'color': 'yellow'})

    questions = [m.question for m in test.mapped_questions.select_related('question').all()]
    hardest_q_text, max_wrong = "अभी पर्याप्त डेटा नहीं", 0
    for i, q in enumerate(questions, 1):
        wrong_count = StudentAnswer.objects.filter(question=q, attempt__test=test).exclude(
            selected_option=q.correct_answer).exclude(selected_option__isnull=True).exclude(selected_option='').count()
        if wrong_count > max_wrong:
            max_wrong, hardest_q_text = wrong_count, f"Q{i}: {q.question_text[:40]}... <br><span style='color:#ff4c4c; font-size:12px;'>({wrong_count} बच्चों ने गलत किया)</span>"

    return JsonResponse({'joined_count': attempts.count(), 'pending_count': pending_students.count(),
                         'hardest_question': hardest_q_text, 'students': students_data})


@login_required
def api_active_students(request, test_id):
    attempts = TestAttempt.objects.filter(test_id=test_id).select_related('student__user')
    return JsonResponse({'count': attempts.count(), 'students': [
        {'name': f"{a.student.user.full_name} ({str(a.student.user.mobile_number)[-4:]})",
         'status': "Submitted 🏁" if a.is_completed else "Live 🟢", 'color': "#ffc107" if a.is_completed else "#00ff00"}
        for a in attempts]})


@login_required
def auto_save_answer(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            attempt = TestAttempt.objects.get(id=data.get('attempt_id'), student=request.user.student_profile)
            question = Question.objects.get(id=data.get('question_id'))
            StudentAnswer.objects.update_or_create(attempt=attempt, question=question,
                                                   defaults={'selected_option': data.get('answer')})
            return JsonResponse({'status': 'success', 'message': 'Answer saved in background'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'error': str(e)})
    return JsonResponse({'status': 'invalid request'})


# ==========================================
# 6. EDITING SINGLE QUESTIONS
# ==========================================
@login_required
def update_correct_answer(request, q_id):
    if request.user.role != 'TEACHER': return redirect('home')
    if request.method == 'POST':
        try:
            mapping = TestQuestionMapping.objects.filter(question_id=q_id,
                                                         test__teacher=request.user.teacher_profile).first()
            if mapping:
                q = mapping.question
                q.correct_answer = request.POST.get('correct_answer')
                q.save()
                messages.success(request, '✅ उत्तर सफलतापूर्वक बदल दिया गया है!')
                return redirect(f"{reverse('preview_test', args=[mapping.test.id])}#question-{q.id}")
        except Exception:
            pass
        messages.error(request, '❌ यह प्रश्न नहीं मिला या कोई एरर आ गया!')
    return redirect(request.META.get('HTTP_REFERER', 'teacher_dashboard'))


@login_required
def upload_question_image(request, q_id):
    if request.user.role != 'TEACHER': return redirect('home')
    if request.method == 'POST' and request.FILES.get('image_file'):
        try:
            mapping = TestQuestionMapping.objects.filter(question_id=q_id,
                                                         test__teacher=request.user.teacher_profile).first()
            if mapping:
                q = mapping.question
                if q.question_image: q.question_image.delete(save=False)
                q.question_image = request.FILES['image_file']
                q.save()
                messages.success(request, '✅ इमेज जुड़ गई है।')
                return redirect(f"{reverse('preview_test', args=[mapping.test.id])}#question-{q.id}")
        except Exception:
            messages.error(request, 'प्रश्न नहीं मिला!')
    return redirect('teacher_dashboard')


@login_required
def delete_question_image(request, q_id):
    if request.user.role != 'TEACHER': return redirect('home')
    try:
        mapping = TestQuestionMapping.objects.filter(question_id=q_id,
                                                     test__teacher=request.user.teacher_profile).first()
        if mapping and mapping.question.question_image:
            mapping.question.question_image.delete(save=True)
            messages.success(request, '✅ इमेज हटा दी गई है।')
            return redirect(f"{reverse('preview_test', args=[mapping.test.id])}#question-{mapping.question.id}")
    except Exception:
        messages.error(request, 'प्रश्न नहीं मिला!')
    return redirect('teacher_dashboard')


@login_required
def delete_single_question(request, q_id):
    if request.user.role != 'TEACHER': return redirect('home')
    try:
        mapping = TestQuestionMapping.objects.filter(question_id=q_id,
                                                     test__teacher=request.user.teacher_profile).first()
        if mapping:
            test_id = mapping.test.id
            mapping.delete()
            messages.success(request, '🗑️ प्रश्न को इस टेस्ट से हटा दिया गया है!')
            return redirect('preview_test', test_id=test_id)
    except Exception:
        pass
    return redirect('teacher_dashboard')


# ==========================================
# 7. PAYMENTS & WALLET
# ==========================================
@login_required
def initiate_payment(request):
    if request.user.role != 'STUDENT': return redirect('home')
    student = request.user.student_profile
    coaching = student.enrolled_coaching
    fee_amount = coaching.subscription_fee if coaching else 99.00
    amount_in_paise = int(fee_amount * 100)
    razorpay_order = razorpay_client.order.create(dict(amount=amount_in_paise, currency="INR", payment_capture='0'))
    PaymentTransaction.objects.create(user=request.user, payment_type='SUBSCRIPTION', amount=fee_amount,
                                      razorpay_order_id=razorpay_order['id'], status='PENDING')
    return render(request, 'portal/payment_page.html',
                  {'razorpay_order_id': razorpay_order['id'], 'razorpay_merchant_key': settings.RAZORPAY_KEY_ID,
                   'razorpay_amount': amount_in_paise, 'currency': "INR", 'fee_amount': fee_amount, 'student': student})


@csrf_exempt
def verify_payment(request):
    if request.method == "POST":
        payment_id, order_id, signature = request.POST.get('razorpay_payment_id', ''), request.POST.get(
            'razorpay_order_id', ''), request.POST.get('razorpay_signature', '')
        try:
            transaction = PaymentTransaction.objects.get(razorpay_order_id=order_id)
        except PaymentTransaction.DoesNotExist:
            return JsonResponse({'status': 'Failed', 'message': 'Transaction not found'})

        try:
            razorpay_client.utility.verify_payment_signature(
                {'razorpay_order_id': order_id, 'razorpay_payment_id': payment_id, 'razorpay_signature': signature})
            transaction.razorpay_payment_id, transaction.razorpay_signature, transaction.status = payment_id, signature, 'SUCCESS'
            transaction.save()

            if transaction.payment_type == 'SUBSCRIPTION':
                student = transaction.user.student_profile
                coaching = student.enrolled_coaching
                validity_days = coaching.subscription_validity_days if coaching else 30
                if student.subscription_expiry_date and student.subscription_expiry_date > timezone.now():
                    student.subscription_expiry_date += timedelta(days=validity_days)
                else:
                    student.subscription_expiry_date = timezone.now() + timedelta(days=validity_days)
                student.save()

                if coaching:
                    teacher_share = transaction.amount - coaching.admin_commission
                    coaching.wallet_balance += teacher_share
                    coaching.save()
                    WalletTransaction.objects.create(user=coaching.user, transaction_type='CREDIT',
                                                     amount=teacher_share,
                                                     description=f"Student {student.display_name} purchased subscription")
                messages.success(request, f"Payment Successful! आपका {validity_days} दिन का प्लान एक्टिवेट हो गया है।")
                return redirect('home')
            elif transaction.payment_type == 'WALLET_RECHARGE':
                profile = transaction.user.student_profile if transaction.user.role == 'STUDENT' else transaction.user.teacher_profile
                profile.wallet_balance += transaction.amount
                profile.save()
                WalletTransaction.objects.create(user=transaction.user, transaction_type='CREDIT',
                                                 amount=transaction.amount, description="Online Wallet Recharge")
                messages.success(request, f"₹{transaction.amount} आपके वॉलेट में सफलता से जुड़ गए हैं!")
                return redirect('wallet_dashboard')
        except razorpay.errors.SignatureVerificationError:
            transaction.status = 'FAILED';
            transaction.save()
            messages.error(request, "Payment Failed or tampered. Please try again.")
    return redirect('home')


@login_required
def pay_via_wallet(request):
    if request.user.role != 'STUDENT': return redirect('home')
    student = request.user.student_profile
    coaching = student.enrolled_coaching
    fee_amount = coaching.subscription_fee if coaching else Decimal('99.00')

    if student.wallet_balance >= fee_amount:
        student.wallet_balance -= fee_amount
        validity_days = coaching.subscription_validity_days if coaching else 30
        if student.subscription_expiry_date and student.subscription_expiry_date > timezone.now():
            student.subscription_expiry_date += timedelta(days=validity_days)
        else:
            student.subscription_expiry_date = timezone.now() + timedelta(days=validity_days)
        student.save()
        WalletTransaction.objects.create(user=request.user, transaction_type='DEBIT', amount=fee_amount,
                                         description=f"Subscription purchased via Wallet")

        if coaching:
            teacher_share = fee_amount - coaching.admin_commission
            coaching.wallet_balance += teacher_share
            coaching.save()
            WalletTransaction.objects.create(user=coaching.user, transaction_type='CREDIT', amount=teacher_share,
                                             description=f"Student {student.display_name} purchased subscription")
        messages.success(request, f"🎉 आपके वॉलेट से ₹{fee_amount} कट गए हैं और प्लान एक्टिवेट हो गया है।")
        return redirect('home')
    else:
        messages.error(request, "❌ आपके वॉलेट में पर्याप्त बैलेंस नहीं है। कृपया ऑनलाइन पेमेंट करें।")
        return redirect('initiate_payment')


@login_required
def wallet_dashboard(request):
    transactions = WalletTransaction.objects.filter(user=request.user).order_by('-created_at')
    context = {'transactions': transactions}
    if request.user.role == 'STUDENT':
        student = request.user.student_profile
        context.update({'balance': student.wallet_balance, 'is_student': True})
        if student.subscription_expiry_date and student.subscription_expiry_date > timezone.now():
            context.update({'plan_status': 'Active (Paid Plan)', 'expiry_date': student.subscription_expiry_date})
        elif student.trial_expiry_date and student.trial_expiry_date > timezone.now():
            context.update({'plan_status': 'Active (Free Trial)', 'expiry_date': student.trial_expiry_date})
        else:
            context.update({'plan_status': 'Expired', 'expiry_date': None})
    elif request.user.role == 'TEACHER':
        context.update({'balance': request.user.teacher_profile.wallet_balance, 'is_teacher': True})
    else:
        return redirect('home')
    return render(request, 'portal/wallet_dashboard.html', context)


@login_required
def add_money(request):
    if request.method == 'POST':
        amount = float(request.POST.get('amount', 0))
        if amount < 10:
            messages.error(request, 'कम से कम ₹10 ऐड करें।')
            return redirect('wallet_dashboard')
        amount_in_paise = int(amount * 100)
        razorpay_order = razorpay_client.order.create(dict(amount=amount_in_paise, currency="INR", payment_capture='0'))
        PaymentTransaction.objects.create(user=request.user, payment_type='WALLET_RECHARGE', amount=amount,
                                          razorpay_order_id=razorpay_order['id'], status='PENDING')
        context = {'razorpay_order_id': razorpay_order['id'], 'razorpay_merchant_key': settings.RAZORPAY_KEY_ID,
                   'razorpay_amount': amount_in_paise, 'currency': "INR", 'fee_amount': amount, 'is_recharge': True}
        if request.user.role == 'STUDENT':
            context['student'] = request.user.student_profile
        elif request.user.role == 'TEACHER':
            context['teacher'] = request.user.teacher_profile
        return render(request, 'portal/payment_page.html', context)
    return redirect('wallet_dashboard')


@login_required
def request_withdrawal(request):
    if request.method == 'POST':
        amount_str, payment_method = request.POST.get('amount', 0), request.POST.get('payment_method')
        payment_details = f"A/C: {request.POST.get('account_number')} | IFSC: {request.POST.get('ifsc_code')}" if payment_method == 'BANK' else request.POST.get(
            'upi_id')
        try:
            amount = Decimal(str(amount_str))
        except:
            return redirect('wallet_dashboard')
        profile = request.user.teacher_profile if request.user.role == 'TEACHER' else request.user.student_profile
        if amount < Decimal('100'): messages.error(request, '❌ कम से कम ₹100 निकाल सकते हैं।'); return redirect(
            'wallet_dashboard')
        if amount > profile.wallet_balance: messages.error(request,
                                                           '❌ आपके वॉलेट में पर्याप्त बैलेंस नहीं है।'); return redirect(
            'wallet_dashboard')

        profile.wallet_balance -= amount
        profile.save()
        WithdrawalRequest.objects.create(user=request.user, amount=amount, payment_method=payment_method,
                                         payment_details=payment_details, status='PENDING')
        WalletTransaction.objects.create(user=request.user, transaction_type='DEBIT', amount=amount,
                                         description=f"Withdrawal Request (Pending)")
        messages.success(request, f'✅ सफलता! आपकी ₹{amount} निकालने की रिक्वेस्ट सबमिट हो गई है।')
    return redirect('wallet_dashboard')


# ==========================================
# 8. UTILS (Profile & Forgot Password)
# ==========================================
def forgot_password_view(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        try:
            user = CustomUser.objects.get(email__iexact=email)
        except CustomUser.DoesNotExist:
            messages.error(request, 'यह ईमेल रजिस्टर नहीं है।')
            return redirect('forgot_password')

        reset_otp = str(random.randint(100000, 999999))
        request.session['reset_email'], request.session['reset_otp'] = user.email, reset_otp

        # 🚀 NAYA: HTML Email Template for Password Reset
        current_time = timezone.localtime(timezone.now()).strftime("%d %b %Y, %I:%M %p")
        subject = f'Mock Test Portal Password Reset OTP: {reset_otp}'
        plain_message = f'आपका पासवर्ड रीसेट OTP है: {reset_otp}'

        html_message = f"""
        <div style="font-family: Arial, sans-serif; background-color: #f4ebdd; padding: 30px 10px;">
            <div style="max-width: 450px; margin: auto; background-color: #ffffff; padding: 30px; border-radius: 12px; border: 1px solid #ddd; box-shadow: 0 4px 10px rgba(0,0,0,0.05);">
                <h2 style="text-align: center; color: #00c6ff; margin-top: 0; font-size: 24px;">Mock Test Portal 🔒</h2>
                <p style="color: #333; font-size: 16px;">नमस्ते <b>{user.full_name}</b>,</p>
                <p style="color: #555; font-size: 15px; line-height: 1.5;">आपके अकाउंट का पासवर्ड रीसेट करने का अनुरोध प्राप्त हुआ है। आपका OTP नीचे दिया गया है:</p>

                <div style="background-color: #4ade80; color: #000; font-size: 36px; font-weight: bold; letter-spacing: 6px; text-align: center; padding: 15px; border-radius: 8px; margin: 25px 0;">
                    {reset_otp}
                </div>

                <p style="font-size: 13px; color: #666; line-height: 1.6;">यह OTP <b>{current_time}</b> पर भेजा गया है और केवल <b>10 मिनट</b> के लिए मान्य है। यदि आपने यह अनुरोध नहीं किया है, तो कृपया इस ईमेल को अनदेखा करें।</p>

                <hr style="border: none; border-top: 1px solid #eee; margin: 25px 0;">
                <p style="font-size: 14px; color: #888; text-align: center;">धन्यवाद,<br><b style="color: #333;">Mock Test Portal Team</b></p>
            </div>
        </div>
        """

        # 🛡️ 🚀 NAYA: Try-Except Block (Crash Protection)
        try:
            # ईमेल भेजने की कोशिश करें
            send_mail(
                subject,
                plain_message,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
                html_message=html_message
            )
            messages.success(request, 'आपके ईमेल पर OTP भेजा गया है।')
            return redirect('verify_reset_otp')

        except Exception as e:
            # 🛑 अगर ईमेल फेल हो जाए, तो साइट क्रैश नहीं होगी, बल्कि असल एरर स्क्रीन पर प्रिंट कर देगी!
            messages.error(request, f'⚠️ ईमेल एरर: {str(e)}')
            return redirect('forgot_password')

    return render(request, 'portal/forgot_password.html')



def verify_reset_otp_view(request):
    if 'reset_email' not in request.session: return redirect('forgot_password')
    if request.method == 'POST':
        if request.POST.get('otp') == request.session.get('reset_otp'):
            request.session['reset_otp_verified'] = True
            messages.success(request, 'OTP वेरीफाई हो गया! नया पासवर्ड बनाएं।')
            return redirect('reset_new_password')
        else:
            messages.error(request, 'OTP गलत है!')
    return render(request, 'portal/verify_reset_otp.html')


def reset_new_password_view(request):
    if not request.session.get('reset_otp_verified'): return redirect('forgot_password')
    if request.method == 'POST':
        if request.POST.get('new_password') != request.POST.get('confirm_password'):
            messages.error(request, 'पासवर्ड मेल नहीं खा रहे हैं।')
            return redirect('reset_new_password')
        user = CustomUser.objects.get(email=request.session.get('reset_email'))
        user.set_password(request.POST.get('new_password'))
        user.save()
        del request.session['reset_email'], request.session['reset_otp'], request.session['reset_otp_verified']
        messages.success(request, '🎉 पासवर्ड सफलतापूर्वक बदल गया है।')
        return redirect('login')
    return render(request, 'portal/set_new_password.html')


@login_required
def profile_view(request):
    if request.method == 'POST':
        new_name = request.POST.get('full_name')
        if new_name:
            request.user.full_name = new_name
            request.user.save()
            if request.user.role == 'STUDENT':
                request.user.student_profile.display_name = f"{new_name.strip().upper()}_{str(request.user.mobile_number)[-4:] if request.user.mobile_number else '0000'}"
                request.user.student_profile.save()

        if request.user.role == 'TEACHER' and request.POST.get('coaching_name'):
            request.user.teacher_profile.coaching_name = request.POST.get('coaching_name')
            request.user.teacher_profile.save()

        messages.success(request, '✅ प्रोफाइल सफलतापूर्वक अपडेट हो गई!')

        # 🚀 NAYA: Save होते ही सीधे डैशबोर्ड पर भेजने का लॉजिक
        if request.user.role == 'TEACHER':
            return redirect('teacher_dashboard')
        elif request.user.is_superuser or request.user.is_staff:
            return redirect('upload_global_bank')
        else:
            return redirect('home')  # Student के लिए

    return render(request, 'portal/profile.html', {'user': request.user})



@login_required
def upload_private_bank(request):
    if request.user.role != 'TEACHER': return redirect('home')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        chapter_id = request.POST.get('chapter_id')
        excel_file = request.FILES['excel_file']

        try:
            chapter = ChapterCategory.objects.get(id=chapter_id)
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue

                Question.objects.create(
                    chapter=chapter,
                    assigned_coaching=request.user.teacher_profile,
                    question_text=str(row[0]), option_a=str(row[1]), option_b=str(row[2]),
                    option_c=str(row[3]), option_d=str(row[4]), correct_answer=str(row[5]).strip().upper(),
                    explanation=str(row[6]) if len(row) > 6 and row[6] else ""
                )
                count += 1

            messages.success(request,
                             f'✅ शानदार! {count} प्रश्न सफलतापूर्वक आपके प्राइवेट बैंक के "{chapter.name}" चैप्टर में जुड़ गए हैं।')
            return redirect('upload_private_bank')

        except Exception as e:
            messages.error(request, f'एरर: {str(e)}')

    exams = ExamCategory.objects.filter(created_by=request.user.teacher_profile)
    return render(request, 'portal/upload_private_bank.html', {'exams': exams})


@login_required
def get_subjects(request):
    exam_id = request.GET.get('exam_id')
    subjects = SubjectCategory.objects.filter(exam_id=exam_id).values('id', 'name')
    return JsonResponse({'subjects': list(subjects)})


@login_required
def get_chapters(request):
    subject_id = request.GET.get('subject_id')
    chapters = ChapterCategory.objects.filter(subject_id=subject_id).values('id', 'name')
    return JsonResponse({'chapters': list(chapters)})


@login_required
def api_delete_private_question(request, q_id):
    if request.method == 'POST':
        try:
            q = Question.objects.get(id=q_id, assigned_coaching=request.user.teacher_profile)
            q.delete()
            return JsonResponse({'status': 'success', 'message': '🗑️ प्रश्न सफलतापूर्वक डिलीट हो गया!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid Request'})


@login_required
def api_move_private_question(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            q_id = data.get('q_id')
            new_chapter_id = data.get('chapter_id')

            q = Question.objects.get(id=q_id, assigned_coaching=request.user.teacher_profile)
            new_chap = ChapterCategory.objects.get(id=new_chapter_id)

            q.chapter = new_chap
            q.save()

            return JsonResponse({'status': 'success', 'message': f'✅ प्रश्न अब "{new_chap.name}" में मूव हो गया है!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    return JsonResponse({'status': 'error', 'message': 'Invalid Request'})


# ==========================================
# 🚀 CREATE NEW CATEGORY VIA AJAX (Teacher & Admin Controlled)
# ==========================================
@login_required
def api_create_category(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            cat_type = data.get('type')
            name = data.get('name').strip()
            parent_id = data.get('parent_id')

            if not name:
                return JsonResponse({'status': 'error', 'message': 'नाम खाली नहीं हो सकता!'})

            teacher = getattr(request.user, 'teacher_profile', None)

            if cat_type == 'exam':
                # 🚀 FIX: Folders are strictly isolated using `created_by=teacher` in lookup
                exam, created = ExamCategory.objects.get_or_create(
                    name__iexact=name, created_by=teacher,
                    defaults={'name': name}
                )
                return JsonResponse({'status': 'success', 'id': exam.id, 'name': exam.name})

            elif cat_type == 'subject':
                if not parent_id: return JsonResponse({'status': 'error', 'message': 'पहले Exam चुनें!'})
                exam = ExamCategory.objects.get(id=parent_id)
                subject, created = SubjectCategory.objects.get_or_create(
                    exam=exam, name__iexact=name, created_by=teacher,
                    defaults={'name': name}
                )
                return JsonResponse({'status': 'success', 'id': subject.id, 'name': subject.name})

            elif cat_type == 'chapter':
                if not parent_id: return JsonResponse({'status': 'error', 'message': 'पहले Subject चुनें!'})
                subject = SubjectCategory.objects.get(id=parent_id)
                chapter, created = ChapterCategory.objects.get_or_create(
                    subject=subject, name__iexact=name, created_by=teacher,
                    defaults={'name': name}
                )
                return JsonResponse({'status': 'success', 'id': chapter.id, 'name': chapter.name})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid Request'})


# ==========================================
# 🌍 UPLOAD TO GLOBAL BANK (ADMIN ONLY)
# ==========================================
@login_required
def upload_global_bank(request):
    if not request.user.is_superuser:
        messages.error(request, '⚠️ Access Denied! केवल एडमिन ही ग्लोबल बैंक में प्रश्न अपलोड कर सकते हैं।')
        return redirect('home')

    if request.method == 'POST' and request.FILES.get('excel_file'):
        chapter_id = request.POST.get('chapter_id')
        excel_file = request.FILES['excel_file']

        try:
            chapter = ChapterCategory.objects.get(id=chapter_id)
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue

                Question.objects.create(
                    chapter=chapter,
                    assigned_coaching=None,
                    question_text=str(row[0]), option_a=str(row[1]), option_b=str(row[2]),
                    option_c=str(row[3]), option_d=str(row[4]), correct_answer=str(row[5]).strip().upper(),
                    explanation=str(row[6]) if len(row) > 6 and row[6] else ""
                )
                count += 1

            messages.success(request,
                             f'🌍 शानदार! {count} प्रश्न सफलतापूर्वक ग्लोबल बैंक के "{chapter.name}" चैप्टर में जुड़ गए हैं।')
            return redirect('upload_global_bank')

        except Exception as e:
            messages.error(request, f'एरर: {str(e)}')

    exams = ExamCategory.objects.filter(created_by__isnull=True)
    return render(request, 'portal/upload_global_bank.html', {'exams': exams})